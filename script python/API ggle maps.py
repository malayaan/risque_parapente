import openpyxl
import requests

def get_elevation(lat, lon, api_key):
    url = f"https://maps.googleapis.com/maps/api/elevation/json?locations={lat},{lon}&key={api_key}"
    response = requests.get(url)
    data = response.json()

    if "results" in data and len(data["results"]) > 0:
        return data["results"][0]["elevation"]
    else:
        return None

def update_excel_file(file_name, api_key):
    # Ouvrir le fichier Excel
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    i=0
    # Parcourir les lignes et obtenir les coordonnées (latitude et longitude)
    for row in range(2, sheet.max_row + 1):
        print(i)
        i+=1
        lat = sheet.cell(row=row, column=1).value
        lon = sheet.cell(row=row, column=2).value

        # Obtenir l'altitude à partir de l'API Google Maps Elevation
        elevation = get_elevation(lat, lon, api_key)

        if elevation is not None:
            # Ajouter l'altitude à la colonne C
            sheet.cell(row=row, column=3).value = elevation
        else:
            sheet.cell(row=row, column=3).value = "na"
            print(f"Erreur lors de la récupération de l'altitude pour les coordonnées ({lat}, {lon})")

    # Enregistrer les modifications dans le fichier Excel
    workbook.save(file_name)

# Remplacez ces valeurs par les vôtres
file_name = "C:/Users/decroux paul/Documents/mission R & D/altitude.xlsx"
api_key = "AIzaSyACetjWQXWhKnsHszg5B6F4U9m43YrrBEM"

update_excel_file(file_name, api_key)
